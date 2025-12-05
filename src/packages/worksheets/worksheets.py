import os
from openpyxl import load_workbook, Workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.drawing.image import Image

from .worksheet_content import WorkSheetContent, Font, Alignment
from .worksheet_styles import WorkSheetStyles


class WorkSheets:

    def __init__(self, path_to_worksheets: str | None = None, title: str | None = None) -> None:
        self.title = title
        self.path_to_workbook = path_to_worksheets
        self.wb = self.__open_workbook(self.path_to_workbook)
        self.ws_style = WorkSheetStyles()
        self.ws_content = WorkSheetContent()

    def set_content_of_active_ws(self) -> None:
        self.__set_rows_height(self.ws_style)
        self.__set_columns_width(self.ws_style)
        self.__merge_cells(self.ws_style.cells_to_merge)
        self.__frame_for_cells(self.ws_style)
        self.__set_text_content(self.ws_content)
        self.__insert_image(self.ws_content)
        self.__set_text_cascade_content(self.ws_content)

    def change_active_ws_title(self, new_ws_title: str) -> None:
        if new_ws_title in self.wb.sheetnames:
            print('Переименовать не получилось, так как такое название листа уже существует в книге!')
        else:
            self.wb.active.title = new_ws_title

    def insert_data_to_sheet(self, data: list[str]) -> None:
        pass

    def _insert_data_in_cell(self, data: str, cell: str, font: Font | None, alignment: Alignment | None) -> None:
        self.wb.active[cell].font = self.ws_content.base_font if font is None else font
        self.wb.active[cell].alignment = self.ws_content.base_alignment if alignment is None else alignment
        self.wb.active[cell] = data

    def save_work_sheet(self) -> None:
        self.wb.save(self.path_to_workbook)

    def create_new_sheet(self, ws_title: str | None) -> None:
        if ws_title is not None and ws_title not in self.wb.sheetnames:
            self.wb.active = self.wb.create_sheet(ws_title)
        else:
            self.wb.active = self.wb.create_sheet()
        self.set_content_of_active_ws()

    def change_active_sheet(self, ws_title) -> None:
        if ws_title in self.wb.sheetnames:
            self.wb.active = self.wb[ws_title]

    def __insert_image(self, content: WorkSheetContent) -> None:
        for cell, image_path in content.image_of_cells.items():
            self.wb.active.add_image(Image(image_path), cell)

    def __set_columns_width(self, ws_style: WorkSheetStyles) -> None:
        for width, columns in ws_style.width_of_columns.items():
            for column in columns:
                self.wb.active.column_dimensions[column].width = width

    def __set_rows_height(self, ws_style: WorkSheetStyles) -> None:
        for height, rows in ws_style.height_of_rows.items():
            for row in rows:
                self.wb.active.row_dimensions[row].height = height

    def __merge_cells(self, range_cells: tuple[str, ...]) -> None:
        for cells in range_cells:
            self.wb.active.merge_cells(cells)

    def __frame_for_cells(self, ws_style) -> None:
        for border, cells_range in ws_style.border_cells.items():
            for row_range in cells_range:
                for row in self.wb.active[row_range]:
                    for cell in row:
                        cell.border = border

    def __set_text_content(self, content: WorkSheetContent) -> None:
        for text, format_of_text in content.text_of_cells.items():
            cell, font, alignment = format_of_text
            self.wb.active[cell] = text
            self.wb.active[cell].font = font
            self.wb.active[cell].alignment = alignment

    def __set_text_cascade_content(self, content: WorkSheetContent) -> None:
        for template, data in content.text_cascade_of_cells.items():
            cells, steps, count, font, alignment = data
            for i in range(count):
                cell = cells['r'] if 'r' in cells else cells['c']
                self.wb.active[cell].font = font
                self.wb.active[cell].alignment = alignment
                self.wb.active[cell] = content.get_text_from_template(template, cells)
                cells = content.change_cells(cells, steps)

    def __open_workbook(self, path: str | None) -> Workbook:
        if path is not None and path.endswith('.xlsx') and os.path.exists(path):
            try:
                wb = load_workbook(path)
            except FileNotFoundError:
                print(f"ОШИБКА: Файл не найден по пути: {os.path.abspath(path)}")
            except InvalidFileException:
                print(f"ОШИБКА: Неверный формат файла. Убедитесь, что это файл .xlsx, а не .xls")
            except Exception as e:
                print(f"Произошла непредвиденная ошибка при открытии файла: {e}")
            else:
                self.title = os.path.basename(path)
                return wb
        return self.__create_new_workbook()

    def __create_new_workbook(self) -> Workbook:
        if self.title is not None:
            self.title += '' if self.title.endswith('.xlsx') else '.xlsx'
        else:
            self.title = 'temp.xlsx'
        self.path_to_workbook = self.title
        return Workbook()
