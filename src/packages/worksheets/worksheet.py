from openpyxl import Workbook

from .worksheet_content import WorkSheetContent
from .worksheet_styles import WorkSheetStyles


class WorkSheet:

    def __init__(self, wb_title: str, ws_title: str) -> None:
        self.wb_title = wb_title
        self.ws_title = ws_title
        self.wb = Workbook()
        self.ws = self.wb.active

    def change_ws_title(self, ws_style: WorkSheetStyles, content: WorkSheetContent) -> None:
        self.ws.title = self.ws_title
        self.__set_rows_height(ws_style)
        self.__set_columns_width(ws_style)
        self.__merge_cells(ws_style.cells_to_merge)
        self.__frame_for_cells(ws_style)
        self.__set_text_content(content)
        self.__insert_image(content)
        self.__set_text_cascade_content(content)

    def save_work_sheet(self):
        self.wb.save(self.wb_title)

    def __insert_image(self, content: WorkSheetContent):
        for cell, image_content in content.image_of_cells.items():
            self.ws.add_image(image_content, cell)

    def __set_columns_width(self, ws_style: WorkSheetStyles) -> None:
        for width, columns in ws_style.width_of_columns.items():
            for column in columns:
                self.ws.column_dimensions[column].width = width

    def __set_rows_height(self, ws_style: WorkSheetStyles) -> None:
        for height, rows in ws_style.height_of_rows.items():
            for row in rows:
                self.ws.row_dimensions[row].height = height

    def __merge_cells(self, range_cells: tuple[str, ...]) -> None:
        for cells in range_cells:
            self.ws.merge_cells(cells)

    def __frame_for_cells(self, ws_style) -> None:
        for border, cells_range in ws_style.border_cells.items():
            for row_range in cells_range:
                for row in self.ws[row_range]:
                    for cell in row:
                        cell.border = border

    def __set_text_content(self, content: WorkSheetContent) -> None:
        for text, format_of_text in content.text_of_cells.items():
            cell, font, alignment = format_of_text
            self.ws[cell] = text
            self.ws[cell].font = font
            self.ws[cell].alignment = alignment

    def __set_text_cascade_content(self, content: WorkSheetContent) -> None:
        for template, data in content.text_cascade_of_cells.items():
            cells, steps, count, font, alignment = data
            for i in range(count):
                cell = cells['r'] if 'r' in cells else cells['c']
                self.ws[cell].font = font
                self.ws[cell].alignment = alignment
                self.ws[cell] = content.get_text_from_template(template, cells)
                cells = content.change_cells(cells, steps)

# if __name__ == '__main__':
#     a = {'c': 'D25', 'r_1': 'D16', 'c_2': 'D24'}
#     b = (1, 1, 1)
#     print(WorkSheetContent.change_cells(a, b))
