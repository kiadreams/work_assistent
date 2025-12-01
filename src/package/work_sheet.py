from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, Side


class WsCellsContent:
    base_font = Font(size=9, name='Arial Cyr')
    base_alignment = Alignment(wrap_text=True)
    text_of_cells = {
        'Ведомость фактического учета трудоемкости работ': (
            'B2',
            Font(bold=True, size=12, name='Arial Cyr'),
            Alignment(horizontal='center', vertical='center')
        ),
        '№ заказа из АСУ ТОиР': ('B4', base_font, base_alignment),
        '№ объекта (его признак из АСУ ТОиР)': ('B5', base_font, base_alignment),
        'Код работы': ('B6', base_font, base_alignment),
        'Наименование работ': ('B7', base_font, base_alignment),
        'Дата начала работ': ('B8', base_font, base_alignment),
        'Дата окончания работ': ('B9', base_font, base_alignment),
        'Наименование МЭС': ('B10', base_font, base_alignment),
        'Наименование ПМЭС': ('B11', base_font, base_alignment),
        'Наименование ПС (ВЛ)': ('B12', base_font, base_alignment),
        'МЭС Юга': ('C10', base_font, base_alignment),
        'Кубанское ПМЭС': ('C11', base_font, base_alignment),
        'ПС 330 кВ Армавир': ('C12', base_font, base_alignment),
        'Доставка персонала (перебазировка)': ('A24', base_font, base_alignment),
        'Итого по бригаде': ('A25', base_font, base_alignment),
        'Итого по бригаде с учетом перебазировки': ('A26', base_font, Alignment(wrap_text=True, vertical='center')),
        'в том числе по категориям рабочих': ('A27', base_font, base_alignment),
        '№ п/п': (
            'A14',
            Font(bold=True, size=10, name='Arial Cyr'),
            Alignment(wrap_text=True, horizontal='center', vertical='center'),
        ),
        'ФИО': (
            'B14',
            Font(bold=True, size=10, name='Arial Cyr'),
            Alignment(wrap_text=True, horizontal='center', vertical='center'),
        ),
        'Разряд': (
            'C14',
            Font(bold=True, size=10, name='Arial Cyr'),
            Alignment(wrap_text=True, horizontal='center', vertical='center'),
        ),
        'Итого, чел/час': (
            'D14',
            Font(bold=True, size=10, name='Arial Cyr'),
            Alignment(wrap_text=True, horizontal='center', vertical='center'),
        ),
    }
    image_of_cells = {
        'F36': Image('../images/img.png'),
    }


class WorkSheetStyle:

    def __init__(self):
        self.width_of_columns = {17: ('B', 'C'), 5: ('A',)}
        self.height_of_rows = {70: (4, 7), 35: (5, 14), 25: (*range(16, 24), *range(28, 36), 26)}
        self.merge_columns_of_row = {
            'C:D': '4', 'E:H': '4', 'C:H': '5:6:7:8:9:10:11:12', 'E:K': '14',
            'A:C': '24:25:26:27', 'B:H': '2', 'B:D': '37', 'H:I': '37'
        }
        self.merge_rows_of_column = {'14:15': 'A:B:C:D'}
        self.merge_rows_and_columns: list[str] = []
        self.border_cells = {
            self.__create_border(side_style='thin', ): ('E16:K23',),
            self.__create_border(side_style='medium'): ('B4:I12', 'A14:D35', 'E14:K15', 'E24:K27'),
            self.__create_border(side_style='medium', border_type='left'): ('L16:L23',),
        }

    @property
    def cells_to_merge(self) -> tuple[str, ...]:
        return tuple(
            self.__get_range_cells_from_rows()
            + self.__get_range_cells_from_columns()
            + self.merge_rows_and_columns
        )

    def __get_range_cells_from_columns(self) -> list[str]:
        range_of_cells = []
        for columns, rows in self.merge_columns_of_row.items():
            column_start, column_end = columns.split(':')
            for row in rows.split(':'):
                 range_of_cells.append(':'.join((column_start + row, column_end + row)))
        return range_of_cells

    def __get_range_cells_from_rows(self) -> list[str]:
        range_of_cells = []
        for rows, columns in self.merge_rows_of_column.items():
            row_start, row_end = rows.split(':')
            for column in columns.split(':'):
                 range_of_cells.append(':'.join((column + row_start, column + row_end)))
        return range_of_cells

    @staticmethod
    def __create_border(side_style: str, border_type='full') -> Border:
        type_side = Side(style=side_style)
        match border_type.lower():
            case 'top':
                return Border(top=type_side)
            case 'left':
                return Border(left=type_side)
            case _:
                return Border(top=type_side, right=type_side, bottom=type_side, left=type_side)


class WorkSheet:

    def __init__(self, wb_title: str, ws_title: str) -> None:
        self.ws_style = WorkSheetStyle()
        self.wb_title = wb_title
        self.ws_title = ws_title
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws_style = WorkSheetStyle()
        self.sign = Image('../images/img.png')

    def change_ws_title(self) -> None:
        self.ws.title = self.ws_title
        self.__set_rows_height(self.ws_style)
        self.__set_columns_width(self.ws_style)
        self.__merge_cells(self.ws_style.cells_to_merge)
        self.__frame_for_cells()
        self.__set_text_content(WsCellsContent())
        self.ws.add_image(self.sign, 'F36')

    def save_work_sheet(self):
        self.wb.save(self.wb_title)

    def __set_columns_width(self, ws_style: WorkSheetStyle) -> None:
        for width, columns in ws_style.width_of_columns.items():
            for column in columns:
                self.ws.column_dimensions[column].width = width

    def __set_rows_height(self, ws_style: WorkSheetStyle) -> None:
        for height, rows in ws_style.height_of_rows.items():
            for row in rows:
                self.ws.row_dimensions[row].height = height

    def __merge_cells(self, range_cells: tuple[str, ...]) -> None:
        for cells in range_cells:
            self.ws.merge_cells(cells)

    def __frame_for_cells(self) -> None:
        for border, cells_range in self.ws_style.border_cells.items():
            for row_range in cells_range:
                for row in self.ws[row_range]:
                    for cell in row:
                        cell.border = border

    def __set_text_content(self, texts: WsCellsContent) -> None:
        for text, format_of_text in texts.text_of_cells.items():
            cell, font, alignment = format_of_text
            self.ws[cell] = text
            self.ws[cell].font = font
            self.ws[cell].alignment = alignment
